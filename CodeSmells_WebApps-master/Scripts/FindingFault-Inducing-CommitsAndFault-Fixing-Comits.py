from pydriller import RepositoryMining, GitRepository
from pprint import pprint
from datetime import date
import xlrd  # to read xls files
import xlwt  # to write to a new xls file
from openpyxl import *  # to write to existing xlsx file
# compute codechurn
from pydriller.metrics.process.code_churn import CodeChurn
import re

repos = GitRepository("/Users/bessghaiernarjess/Documents/PhD_ETS/git/symfony")
# keywords = 29
Keywords = ["fix", "fixed", "fixes", "fixing", "close", "closed", "closes", "closing", "bug", "bugs", "fault", "faults", "error", "errors",
            "defect", "defects", "issue", "issues", "failure", "failures", "crash", "crashes", "except", "fail",
            "resolves",
            "resolved", "solve", "solves", "solved"]

i = 0
i1 = 0
ch = ""
ch1 = ""
NbBugs = 0
NbFiles = 0

fix = 0
fixed = 0
fixes = 0
fixing = 0
close = 0
closed = 0
closes = 0
closing = 0
bug = 0
bugs = 0
error = 0
errors = 0
defect = 0
defects = 0
issue = 0
issues = 0
failure = 0
failures = 0
crash = 0
crashes = 0
exceptX = 0
fail = 0
resolves = 0
resolved = 0
solve = 0
solves = 0
solved = 0
fault=0
faults=0

code_churn_buggyF = 0
nbCommits = 0
BuggyCommits = ""
ChurnFixing = 0
found = 0
List_days = []
Smelly_days = []
Nonsmelly_days = []
statusMatch = 0
statusMatch1 = 0
# FirstCommit="d00a85f1d87718ea929edb264ef0d056e022608d"
# CurrentCommit="8d128091ca631d3f6f683edddba1e89ffbe82998"
FoundKeywords = ""


def words_in_string(word_list, a_string):
    return set(word_list).intersection(a_string.split())


##------------------------------
##------------------------------
# READ ABD WRITE TO DATASETS
##------------------------------
# -------------------------------


path_Keywords = ("/Users/bessghaiernarjess/Documents/PhD_ETS/Contrib2-Fault-proneness/Datasets/KeywordsFrequency.xlsx")
path_RQ1SmellyFiles = (
    "/Users/bessghaiernarjess/Documents/PhD_ETS/Contrib2-Fault-proneness/Datasets/RQ1_CodeChurn_NbBugs/Symphony/RQ1-Symphony-SmellyFiles_CodeChurnBugs.xlsx")
path_RQ1NonSmellyFiles = (
    "/Users/bessghaiernarjess/Documents/PhD_ETS/Contrib2-Fault-proneness/Datasets/RQ1_CodeChurn_NbBugs/Symphony/RQ1-Symphony-NonSmellyFiles_CodeChurnBugs.xlsx")
path_RQ2SmellyFiles = (
    "/Users/bessghaiernarjess/Documents/PhD_ETS/Contrib2-Fault-proneness/Datasets/RQ2_TimeTillBugOcc/symphony/RQ2-symphony-SmellyFiles_CodeChurnBugs.xlsx")
path_RQ2NonSmellyFiles = (
    "/Users/bessghaiernarjess/Documents/PhD_ETS/Contrib2-Fault-proneness/Datasets/RQ2_TimeTillBugOcc/symphony/RQ2-symphony-NonSmellyFiles_CodeChurnBugs.xlsx")
path_RQ3LocSmellyFiles = (
    "/Users/bessghaiernarjess/Documents/PhD_ETS/Contrib2-Fault-proneness/Datasets/RQ3_Loc/symphony/LocChurnSmelly.xlsx")
path_RQ3LocNonSmellyFiles = (
    "/Users/bessghaiernarjess/Documents/PhD_ETS/Contrib2-Fault-proneness/Datasets/RQ3_Loc/symphony/LocChurnNonSmelly.xlsx")

###VARIABLES TO CHANGE
FirstCommit = "66b2e9662c44d478b69e48278aa54079a006eb42"  #
CurrentCommit = ""  #
release = 21  # 4
NbRelease = "release 18"
SmellyDayslength = 31 # 2      RQ1 files must be filled with 0
NonsmellyDayslength = 32 # 2
LocLineSmelly = 2397  # 1
LocLineNonSmelly = 2600  # 1
for Commit in RepositoryMining("/Users/bessghaiernarjess/Documents/PhD_ETS/git/symfony",
                               from_commit="969d709ad428076bf1084e386dc26dd904d9fb84",
                               to_commit="66b2e9662c44d478b69e48278aa54079a006eb42",
                               only_modifications_with_file_types=['.php']).traverse_commits():

    # WE IDENTIFY ALL MODIFIED FILES IN ONE SINGLE FAULT-FIXING COMMITS
    for modified_file in Commit.modifications:

        # for  c in Keywords:
        #  if c.lower() in Commit.msg:
        # CHECK IF KEYWORDS EXIST
        string = Commit.msg.lower()
        word = words_in_string(Keywords, string)
        s = list(word)
        for d1 in RepositoryMining("/Users/bessghaiernarjess/Documents/PhD_ETS/git/symfony",
                                   single=CurrentCommit).traverse_commits():

            for d in RepositoryMining("/Users/bessghaiernarjess/Documents/PhD_ETS/git/symfony",
                                      single=FirstCommit).traverse_commits():
                # print("FirstCommit "+ d.committer_date.strftime("%Y-%m-%d"))
                # AT LEAST ONE KEYWORD EXISTS, WITH PHP EXTENSION AND THE DATE IS SUPERIOR THAN THE FIRST RELEASE
                if len(s) != 0 and ".php" in modified_file.filename and Commit.committer_date > d.committer_date and Commit.committer_date < d1.committer_date:
                    code_churn = int(modified_file.added) + int(modified_file.removed)
                    print(
                        'Hash= {} ** keyword= {} ** file_name= {} ** commit_date= {} LOC= {} Code Churn= {} (+{} , -{})'
                        .format(Commit.hash, s, modified_file.filename, Commit.committer_date.strftime("%Y-%m-%d"),
                                modified_file.nloc, code_churn, modified_file.added, modified_file.removed))
                    NbFiles += 1
                    nbCommits += 1
                    # print(Commit.msg)

                    if "fix" in s: fix += 1
                    if "fixed" in s: fixed += 1
                    if "fixes" in s: fixes += 1
                    if "fixing" in s: fixing += 1
                    if "close" in s: close += 1
                    if "closed" in s: closed += 1
                    if "closes" in s: closes += 1
                    if "closing" in s: closing += 1
                    if "bug" in s: bug += 1
                    if "bugs" in s: bugs += 1
                    if "error" in s: error += 1
                    if "errors" in s: errors += 1
                    if "defect" in s: defect += 1
                    if "defects" in s: defects += 1
                    if "issue" in s: issue += 1
                    if "issues" in s: issues += 1
                    if "failure" in s: failure += 1
                    if "failures" in s: failures += 1
                    if "crash" in s: crash += 1
                    if "crashes" in s: crashes += 1
                    if "fail" in s: fail += 1
                    if "fault" in s: fault += 1
                    if "faults" in s: faults += 1
                    if "resolves" in s: resolves += 1
                    if "resolved" in s: resolved += 1
                    if "solve" in s: solve += 1
                    if "solves" in s: solves += 1
                    if "solved" in s: solved += 1
                    if "except" in s: exceptX += 1
                    #
                    # print(Commit.msg)
                    # get files code modifications
                    # == displays all code modifications
                    # print(modified_file.diff)
                    # to facilitate the parsing of code modifications, we use the GitRepository Class in line 3
                    # diff =modified_file.diff
                    # print(diff)
                    # parsed_diff= repos.parse_diff(diff)
                    # pprint(parsed_diff)

                    # GET ALL BUGGY COMMITS
                    bug_inducing_commits = repos.get_commits_last_modified_lines(Commit, modified_file)
                    # print(bug_inducing_commits)
                    BuggyCommits = bug_inducing_commits
                    # print(BuggyCommits)

                    # x = re.search("(?<=\')(.*?)(?=\')", str(BuggyCommits))
                    # print(x)
                    # match=re.findall("(?<=\')(.*?)(?=\')", str(BuggyCommits))
                    match = re.findall("(?<=\ ')[a-zA-Z0-9 ]*", str(BuggyCommits))
                    # print(match)
                    match1 = re.findall("(?<=\ {')[a-zA-Z0-9 ]*", str(BuggyCommits))
                    # print(match1)

                    NbBugs = len(match) + len(match1)

                    #print("Number of buggy commits before and after introduction of smells= " + str(NbBugs))
                    #             for i in match:
                    #                print(i)
                    #             for i1 in match1:
                    #                print(i1)
                    #           #EXTRACT ALL BUGGY COMMITS FROM THE LISTS
                    if NbBugs != 0:
                        nbCommits += NbBugs
                        ##GET buggy commits date and difference with the fixing commits
                        # for d in RepositoryMining("/Users/bessghaiernarjess/Documents/PhD_ETS/git/joomla-cms",single=FirstCommit).traverse_commits():
                        #print("FirstCommit " + d.committer_date.strftime("%Y-%m-%d"))
                        #print("CurrentCommit " + d1.committer_date.strftime("%Y-%m-%d"))
                        for i in match:
                            for c in RepositoryMining("/Users/bessghaiernarjess/Documents/PhD_ETS/git/symfony",
                                                      single=i).traverse_commits():

                                f_date = date(int(d.committer_date.strftime("%Y")),
                                              int(d.committer_date.strftime("%m")),
                                              int(d.committer_date.strftime("%d")))
                                l_date = date(int(c.committer_date.strftime("%Y")),
                                              int(c.committer_date.strftime("%m")),
                                              int(c.committer_date.strftime("%d")))
                                days = l_date - f_date
                                # print(days.days)
                                for buggy_file in c.modifications:
                                    # EXTRACT THE CODE CHURN OF TARGETED FILE MODIFIED BY THE BUGGY COMMIT
                                    if not (modified_file.old_path is None):
                                        if buggy_file.old_path == modified_file.old_path:
                                            code_churn_buggy = int(buggy_file.added) + int(buggy_file.removed)
                                            statusMatch = 1
                                            # LOC=int(modified_file.nloc)+int(buggy_file.added)-int(buggy_file.removed)

                                            # MAKE SURE WE ONLY EXTRACT BUGS INTRODUCED AFTER THE INTRODUCION OF SMELLS
                                            if c.committer_date > d.committer_date:
                                                bugs += 1
                                                # print(
                                                #     '{} {}  {} days_till_first_bug_occurrence={} code_churn= {}'.format(
                                                #         c.hash, buggy_file.filename,
                                                #         c.committer_date.strftime("%Y-%m-%d"), days.days,
                                                #         code_churn_buggy))
                                                List_days.append(days.days)

                                            # code_churn_buggyF=code_churn_buggyF+code_churn_buggy
                        # print(i.author_date.strftime("%Y-%m-%d"))

                        for i1 in match1:
                            for c1 in RepositoryMining("/Users/bessghaiernarjess/Documents/PhD_ETS/git/symfony",
                                                       single=i1).traverse_commits():

                                f_date = date(int(d.committer_date.strftime("%Y")),
                                              int(d.committer_date.strftime("%m")),
                                              int(d.committer_date.strftime("%d")))
                                l_date = date(int(c1.committer_date.strftime("%Y")),
                                              int(c1.committer_date.strftime("%m")),
                                              int(c1.committer_date.strftime("%d")))
                                days = l_date - f_date
                                # print(days.days)
                                for buggy_file in c1.modifications:
                                    if not (modified_file.old_path is None):
                                        if buggy_file.old_path == modified_file.old_path:
                                            code_churn_buggy1 = int(buggy_file.added) + int(buggy_file.removed)
                                            statusMatch1 = 1
                                            # LOC=int(modified_file.nloc)+int(buggy_file1.added)-int(buggy_file1.removed)
                                            if c1.committer_date > d.committer_date:
                                                bugs += 1
                                                # print(
                                                #     '{} {}  {} days_till_first_bug_occurrence= {} code_churn= {}'.format(
                                                #         c1.hash, buggy_file.filename,
                                                #         c1.committer_date.strftime("%Y-%m-%d"), days.days,
                                                #         code_churn_buggy1))
                                                List_days.append(days.days)

                                            # code_churn_buggyF=code_churn_buggyF+code_churn_buggy1
                        # print(i.author_date.strftime("%Y-%m-%d"))

                    ##WRITE TO RQ1Smelly.XLS

                    wb = load_workbook(path_RQ1SmellyFiles)
                    ws = wb["Sheet1"]

                    # check the file exists

                    for i in range(ws.max_row):
                        # print(ws.max_row)
                        wcell1 = ws.cell(i + 2, 1)
                        if not (modified_file.old_path is None):
                            if str(wcell1.value) in modified_file.old_path:
                                LocLineSmelly += 1
                                wb7 = load_workbook(path_RQ3LocSmellyFiles)
                                ws7 = wb7["Sheet1"]
                                wcell1 = ws7.cell(LocLineSmelly, 1)
                                wcell1.value = NbRelease
                                wcell1 = ws7.cell(LocLineSmelly, 2)
                                wcell1.value = modified_file.nloc
                                wcell1 = ws7.cell(LocLineSmelly, 3)
                                wcell1.value = code_churn
                                wcell1 = ws7.cell(LocLineSmelly, 4)
                                wcell1.value = modified_file.added
                                wcell1 = ws7.cell(LocLineSmelly, 5)
                                wcell1.value = modified_file.removed

                                wb7.save(path_RQ3LocSmellyFiles)

                                found = 1

                                wcell1 = ws.cell(i + 2, 2)
                                if NbBugs != 0:
                                    wcell1.value = wcell1.value + NbBugs
                                    # print (str(NbBugs))
                                wcell1 = ws.cell(i + 2, 3)

                                if bugs != 0:
                                    wcell1.value = wcell1.value + bugs
                                    # print (str(bugs))
                                    wcell2 = ws.cell(i + 2, release)
                                    if statusMatch1 == 1 and statusMatch == 1:
                                        wcell2.value = wcell1.value + int(code_churn_buggy + code_churn_buggy1)
                                    if statusMatch1 == 1 and statusMatch == 0:
                                        wcell2.value = wcell1.value + int(0 + code_churn_buggy1)
                                    if statusMatch1 == 0 and statusMatch == 1:
                                        wcell2.value = wcell1.value + int(code_churn_buggy + 0)

                                        ## DEAL WITH RQ2 DATASET

                                    wb3 = load_workbook(path_RQ2SmellyFiles)
                                    ws3 = wb3["Sheet1"]
                                    Smelly_days.append(len(List_days))
                                    wcell1 = ws3.cell(1, SmellyDayslength)
                                    wcell1.value = NbRelease
                                    for i1 in range(len(List_days)):
                                        wcell1 = ws3.cell(i + 2, i1 + SmellyDayslength)
                                        wcell1.value = List_days[i1]
                                    wb3.save(path_RQ2SmellyFiles)

                                    ## write code churn in keywords.xlsx
                                    wb5 = load_workbook(path_Keywords)
                                    ws5 = wb5["Sheet1"]
                                    wcell1 = ws5.cell(2, 5)
                                    wcell1.value = wcell1.value + code_churn
                                    wb5.save(path_Keywords)

                                    wb7 = load_workbook(path_RQ3LocSmellyFiles)
                                    ws7 = wb7["Sheet1"]
                                    wcell1 = ws7.cell(LocLineSmelly, 6)
                                    wcell1.value = buggy_file.added
                                    wcell1 = ws7.cell(LocLineSmelly, 7)
                                    wcell1.value = buggy_file.removed
                                    wb7.save(path_RQ3LocSmellyFiles)
                    wb.save(path_RQ1SmellyFiles)

                    ##WRITE TO RQ1NonSmelly.XLS
                    if found != 1:

                        wb1 = load_workbook(path_RQ1NonSmellyFiles)
                        ws1 = wb1["Sheet1"]

                        # check the file exists

                        for i in range(ws1.max_row):
                            wcell1 = ws1.cell(i + 2, 1)
                            if not (modified_file.old_path is None):
                                if str(wcell1.value) in modified_file.old_path:
                                    found = 0
                                    LocLineNonSmelly += 1
                                    wb8 = load_workbook(path_RQ3LocNonSmellyFiles)
                                    ws8 = wb8["Sheet1"]
                                    wcell1 = ws8.cell(LocLineNonSmelly, 1)
                                    wcell1.value = NbRelease
                                    wcell1 = ws8.cell(LocLineNonSmelly, 2)
                                    wcell1.value = modified_file.nloc
                                    wcell1 = ws8.cell(LocLineNonSmelly, 3)
                                    wcell1.value = code_churn
                                    wcell1 = ws8.cell(LocLineNonSmelly, 4)
                                    wcell1.value = modified_file.added
                                    wcell1 = ws8.cell(LocLineNonSmelly, 5)
                                    wcell1.value = modified_file.removed

                                    wb8.save(path_RQ3LocNonSmellyFiles)

                                    wcell1 = ws1.cell(i + 2, 2)
                                    if NbBugs != 0:
                                        wcell1.value = wcell1.value + NbBugs
                                        # print (str(NbBugs))
                                    wcell1 = ws1.cell(i + 2, 3)

                                    if bugs != 0:
                                        wcell1.value = wcell1.value + bugs
                                        # print (str(bugs))
                                        wcell2 = ws1.cell(i + 2, release)
                                        if statusMatch1 == 1 and statusMatch == 1:
                                            wcell2.value = wcell1.value + int(code_churn_buggy + code_churn_buggy1)
                                        if statusMatch1 == 1 and statusMatch == 0:
                                            wcell2.value = wcell1.value + int(0 + code_churn_buggy1)
                                        if statusMatch1 == 0 and statusMatch == 1:
                                            wcell2.value = wcell1.value + int(code_churn_buggy + 0)
                                            # print (str(code_churn_buggy+code_churn_buggy1))

                                        ## DEAL WITH RQ2 DATASET

                                        wb4 = load_workbook(path_RQ2NonSmellyFiles)
                                        ws4 = wb4["Sheet1"]
                                        Nonsmelly_days.append(len(List_days))
                                        wcell1 = ws4.cell(1, NonsmellyDayslength)
                                        wcell1.value = NbRelease
                                        for i1 in range(len(List_days)):
                                            wcell1 = ws4.cell(i + 2, i1 + NonsmellyDayslength)
                                            wcell1.value = List_days[i1]
                                        wb4.save(path_RQ2NonSmellyFiles)
                                        ## write code churn in keywords.xlsx
                                        wb6 = load_workbook(path_Keywords)
                                        ws6 = wb6["Sheet1"]
                                        wcell1 = ws6.cell(2, 6)
                                        wcell1.value = wcell1.value + code_churn
                                        wb6.save(path_Keywords)

                                        wb8 = load_workbook(path_RQ3LocNonSmellyFiles)
                                        ws8 = wb8["Sheet1"]
                                        wcell1 = ws8.cell(LocLineNonSmelly, 6)
                                        wcell1.value = buggy_file.added
                                        wcell1 = ws8.cell(LocLineNonSmelly, 7)
                                        wcell1.value = buggy_file.removed
                                        wb8.save(path_RQ3LocNonSmellyFiles)
                        wb1.save(path_RQ1NonSmellyFiles)

                    print("-----------------------------------------")
                    #print("-----------------------------------------")
                    code_churn_buggyF = 0
                    FoundKeywords = ""
                    ChurnFixing = 0
                    NbBugs = 0
                    bugs = 0
                    found = 0
                    List_days = []
                # Count the frequency of keywrods use

print("-----------------------------------------")
print("------------------FINISH-----------------")
print("-----------------------------------------")
# print("NbFiles= "+str(NbFiles))
# print("nbCommits= "+str(nbCommits))
# print("fix= "+str(fix))
# print("fixed= "+str(fixed))
# print("fixes= "+str(fixes))
# print("fixing= "+str(fixing))
# print("close= "+str(close))
# print("closed= "+str(closed))
# print("closes=" +str(closes))
# print("closing=" +str(closing))
# print("bug= "+str(bug))
# print("bugs= "+str(bugs))
# print("error= "+str(error))
# print("errors= "+str(errors))
# print("defect= "+str(defect))
# print("defects= "+str(defects))
# print("issue= "+str(issue))
# print("issues= "+str(issues))
# print("failure= "+str(failure))
# print("failures= "+str(failures))
# print("crash= "+str(crash))
# print("crashes= "+str(crashes))
# print("resolves= "+str(resolves))
# print("resolved= "+str(resolved))
# print("fail= "+str(fail))
# print("fault= "+str(fault))
# print("faults= "+str(faults))
# print("solve= "+str(solve))
# print("solved= "+str(solved))
# print("solves= "+str(solves))
# print("except= "+str(exceptX))

###find max in two lists
if len(Smelly_days) > 0:
    print("Smelly Largest days:", max(Smelly_days))
else:
    print("Smelly Largest days:", 0)
if len(Nonsmelly_days) > 0:
    print("Non Smelly largest days:", max(Nonsmelly_days))
else:
    print("Non Smelly Largest days:", 0)
print("LocLineSmelly", LocLineSmelly)
print("LocLineNonSmelly", LocLineNonSmelly)

keys = [fix, fixed, fixes, fixing, close, closed, closes, closing, bug, bugs, fault, faults, error, errors, defect, defects, issue,
        issues, failure, failures, crash, crashes, resolves,
        resolved, fail, solve, solved, solves,  exceptX]

##WRITE TO KEYWORDS.XLS

wb2 = load_workbook(path_Keywords)
ws2 = wb2["Sheet1"]
for i in range(len(keys)):
    wcell1 = ws2.cell(i + 2, 2)
    wcell1.value = wcell1.value + keys[i]
wcell1 = ws2.cell(2, 3)
wcell1.value = wcell1.value + nbCommits
wcell1 = ws2.cell(2, 4)
wcell1.value = wcell1.value + NbFiles
wb2.save(path_Keywords)

Smelly_days = []
Nonsmelly_days = []
