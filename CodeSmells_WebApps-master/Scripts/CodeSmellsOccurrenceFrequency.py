'''
Created on Jun 29, 2020

@author: NarjesBessghaier
'''

import pandas as pd
import csv
from openpyxl import *
import openpyxl
from pathlib import Path
import xlrd 
import six
#store_data = pd.read_csv('/Users/bessghaiernarjess/Documents/workspace-luna/fault-inducing-commits/Co-occurrence/store_data.csv', header=None)
loc = ("/Users/bessghaiernarjess/Documents/PhD_ETS/Contrib2-Fault-proneness/Occurrence/OccAllApps.xlsx") 


#COMPUTE THE FREQUENCY OF SINGLE ITEMS
CC= 0;
NPC= 0;
EML= 0;
ECL= 0;
EPL= 0;
TMPM= 0;
TMM= 0;

CBO= 0;
    
length=107228
cols=7

ECL_TMPM=0;
TMM_ECL=0;
HMC_TMM=0;
TMM_EPL=0
ECL_EML=0
TMPM_EML=0
EPL_NPC=0
TMM_NPC=0
CBO_EPL=0
TMM_EML=0
wb=load_workbook(loc)           
ws=wb["Sheet1"] 
print('LOC='+ str(ws.max_row))
for i in range(1,ws.max_row):
    for j in range (1,10):
        wcell1=ws.cell(i,j)  
        if str(wcell1.value)=="Cyclomatic Complexity":
            CC=1
        if str(wcell1.value)=="Excessive Class Length":
            ECL=1  
            #print(ECL)
        if str(wcell1.value)=="Too Many Public Methods":
            TMPM=1 
            #print(TMPM)
        if str(wcell1.value)=="Too Many Methods":
            TMM=1
        if str(wcell1.value)=="Excessive Method Length":
            EML=1
        if str(wcell1.value)=="Excessive Parameter List":
            EPL=1   
  
        if str(wcell1.value)=="NPath Complexity":
            NPC=1 
        if str(wcell1.value)=="Coupling Between Objects":
            CBO=1
            
    if CC==1 and TMM ==1 :
         HMC_TMM+=1;
         print(i)
    if ECL==1 and TMM ==1 :
         TMM_ECL+=1;
         
    if ECL==1 and TMPM ==1 :
         ECL_TMPM+=1;     
         
         
    if ECL==1 and EML ==1 :
         ECL_EML+=1; 
         
    if TMPM==1 and EML ==1 :
         TMPM_EML+=1; 
         
    if EPL==1 and NPC ==1 :
         EPL_NPC+=1;     
         
    if TMM==1 and NPC ==1 :
         TMM_NPC+=1; 
       
    if CBO==1 and EPL ==1 :
         CBO_EPL+=1;
         
    if TMM==1 and EML ==1 :
         TMM_EML+=1;   
         
    if TMM==1 and EPL ==1 :
         TMM_EPL+=1;               
                 
    CC= 0;
    NPC= 0;
    EML= 0;
    ECL= 0;
    EPL= 0;
    TMPM= 0;
    TMM= 0;

    CBO= 0;
 
print("HMC_TMM= "+str(HMC_TMM))
print("TMM_ECL= "+str(TMM_ECL))
print("ECL_TMPM= "+str(ECL_TMPM))
print("TMM_EPL= "+str(TMM_EPL))
print("ECL_EML= "+str(ECL_EML))
print("TMPM_EML= "+str(TMPM_EML))
print("EPL_NPC= "+str(EPL_NPC))
print("TMM_NPC= "+str(TMM_NPC))
print("CBO_EPL= "+str(CBO_EPL))
print("TMM_EML= "+str(TMM_EML))


print("HMC_TMM= "+str((HMC_TMM*100)/length))
print("TMM_ECL= "+str((TMM_ECL*100)/length))
print("ECL_TMPM= "+str((ECL_TMPM*100)/length))      
print("TMM_EPL= "+str((TMM_EPL*100)/length))
print("ECL_EML= "+str((ECL_EML*100)/length))
print("TMPM_EML= "+str((TMPM_EML*100)/length))
print("EPL_NPC= "+str((EPL_NPC*100)/length))
print("TMM_NPC= "+str((TMM_NPC*100)/length))
print("CBO_EPL= "+str((CBO_EPL*100)/length))
print("TMM_EML= "+str((TMM_EML*100)/length))


